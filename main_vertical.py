import customtkinter as ctk
from tkinter import messagebox
from datetime import datetime
import os
import sys  # EXE yolunu bulmak için eklendi
import math
import openpyxl
from openpyxl import Workbook, load_workbook

# --- AYARLAR ---
ctk.set_appearance_mode("Light")
ctk.set_default_color_theme("blue")

class TPNApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("TPN Hesaplayıcı ve Kayıt Sistemi")
        self.geometry("1100x850")
        
        self.latest_results = None
        
        # --- EXE YOLUNU BULMA (Permission Denied Çözümü) ---
        if getattr(sys, 'frozen', False):
            # Eğer EXE olarak çalışıyorsa
            self.app_path = os.path.dirname(sys.executable)
        else:
            # Eğer Python kodu olarak çalışıyorsa
            self.app_path = os.path.dirname(os.path.abspath(__file__))

        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        # --- SOL TARAFA: GİRİŞ PANELİ ---
        self.input_frame = ctk.CTkScrollableFrame(self, label_text="Veri Girişi")
        self.input_frame.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

        self.entries = {}
        
        # 1. Hasta Bilgileri
        self.add_section_header(self.input_frame, "HASTA BİLGİLERİ")
        self.add_input("Adı Soyadı", "text")
        self.add_input("Tarih", "date")
        self.add_input("Bebek Gramı (gr)", "number")

        # 2. Hedefler
        self.add_section_header(self.input_frame, "HEDEFLER")
        self.add_input("Hacim Katsayısı (cc/kg)", "number")
        self.add_input("Glukoz Hedef (mg/kg/dk)", "number")

        # 3. Katsayılar
        self.add_section_header(self.input_frame, "KATSAYILAR")
        self.add_input("NaCl %3 Katsayı", "number")
        self.add_input("KCL Katsayısı", "number")
        self.add_input("KPO4 Katsayısı", "number")
        self.add_input("ADDAMEL Katsayısı", "number")
        self.add_input("CERNEVİT (1=Var)", "number")
        self.add_input("LİPİT Katsayısı", "number")
        self.add_input("DİPEPTİVEN Katsayısı", "number")
        self.add_input("OMEGAVEN Katsayısı", "number")
        self.add_input("PRIMENE Katsayısı", "number")
        self.add_input("MgSO4 Katsayısı", "number")
        self.add_input("Ca Katsayısı", "number")
        self.add_input("Heparin Katsayısı", "number")

        # Butonlar
        self.btn_frame = ctk.CTkFrame(self.input_frame, fg_color="transparent")
        self.btn_frame.pack(pady=20, fill="x")

        self.btn_calc = ctk.CTkButton(self.btn_frame, text="HESAPLA", command=self.calculate, height=40, fg_color="green", font=("Arial", 14, "bold"))
        self.btn_calc.pack(side="left", expand=True, padx=5)

        self.btn_save = ctk.CTkButton(self.btn_frame, text="EXCEL'E KAYDET", command=self.save_to_excel, height=40, fg_color="#1f6aa5", state="disabled", font=("Arial", 14, "bold"))
        self.btn_save.pack(side="right", expand=True, padx=5)

        # --- SAĞ TARAF: SONUÇ PANELİ ---
        self.result_frame = ctk.CTkFrame(self)
        self.result_frame.grid(row=0, column=1, padx=10, pady=10, sticky="nsew")
        
        lbl_res = ctk.CTkLabel(self.result_frame, text="SONUÇ TABLOSU", font=("Arial", 20, "bold"))
        lbl_res.pack(pady=15)

        self.result_labels = {}
        self.output_keys = [
            "ADI SOYADI", "TOTAL", "Tarih", "NaCl %0.9", "%3 NaCl (serum sale)",
            "KCL", "KPO4", "ADDAMEL", "CERNEVİT", "LİPİT", 
            "DİPEPTİVEN", "OMEGAVEN", "PRIMENE", 
            "DEXTROZ 5", "DEXTROZ 10", "DEXTROZ 20", "DEXTROZ 30", "DEXTROZ 50",
            "MgSO4", "Ca", "HEPARİN"
        ]

        self.res_container = ctk.CTkScrollableFrame(self.result_frame, fg_color="transparent")
        self.res_container.pack(fill="both", expand=True, padx=5)

        for key in self.output_keys:
            row_frame = ctk.CTkFrame(self.res_container, fg_color=("gray90", "gray30"))
            row_frame.pack(fill="x", pady=2)
            
            lbl_key = ctk.CTkLabel(row_frame, text=key, anchor="w", width=180, font=("Arial", 13, "bold"))
            lbl_key.pack(side="left", padx=10)
            
            lbl_val = ctk.CTkLabel(row_frame, text="-", font=("Arial", 13))
            lbl_val.pack(side="right", padx=10)
            
            self.result_labels[key] = lbl_val

    def add_section_header(self, parent, text):
        lbl = ctk.CTkLabel(parent, text=text, font=("Arial", 12, "bold"), text_color="gray")
        lbl.pack(fill="x", pady=(15, 5), padx=5)

    def add_input(self, label_text, input_type):
        frame = ctk.CTkFrame(self.input_frame, fg_color="transparent")
        frame.pack(fill="x", pady=2)
        
        lbl = ctk.CTkLabel(frame, text=label_text, width=180, anchor="w")
        lbl.pack(side="left", padx=5)
        
        entry = ctk.CTkEntry(frame)
        entry.pack(side="left", fill="x", expand=True, padx=5)
        
        if input_type == "date":
            entry.insert(0, datetime.now().strftime("%d/%m/%Y"))
        
        # --- SIFIR GİRİLİ GELME ÖZELLİĞİ ---
        if input_type == "number":
            entry.insert(0, "0")
            
        self.entries[label_text] = entry

    def get_val(self, key):
        try:
            txt = self.entries[key].get().replace(',', '.')
            if not txt: return 0.0 # Boşsa 0 dön
            return float(txt)
        except ValueError:
            return 0.0
            
    def get_text(self, key):
        return self.entries[key].get()

    def calculate(self):
        try:
            weight_gr = self.get_val("Bebek Gramı (gr)")
            weight_kg = weight_gr / 1000.0
            
            if weight_kg <= 0:
                messagebox.showerror("Hata", "Bebek ağırlığı giriniz.")
                return

            vol_coeff = self.get_val("Hacim Katsayısı (cc/kg)")
            target_glucose = self.get_val("Glukoz Hedef (mg/kg/dk)")

            # --- HACİM HESAPLAMA ---
            primene_vol = self.get_val("PRIMENE Katsayısı") * weight_kg * 10
            lipid_vol = self.get_val("LİPİT Katsayısı") * weight_kg * 2
            nacl3_vol = self.get_val("NaCl %3 Katsayı") * weight_kg
            kcl_vol = self.get_val("KCL Katsayısı") * weight_kg
            kpo4_vol = self.get_val("KPO4 Katsayısı") * weight_kg
            addamel_vol = self.get_val("ADDAMEL Katsayısı") * weight_kg
            
            # Cernevit üste yuvarla
            raw_cernevit = self.get_val("CERNEVİT (1=Var)") * weight_kg 
            cernevit_vol = math.ceil(raw_cernevit) 
            
            dipeptiven_vol = self.get_val("DİPEPTİVEN Katsayısı") * weight_kg
            omegaven_vol = self.get_val("OMEGAVEN Katsayısı") * weight_kg
            mgso4_vol = self.get_val("MgSO4 Katsayısı") * weight_kg
            ca_vol = self.get_val("Ca Katsayısı") * weight_kg
            heparin_vol = self.get_val("Heparin Katsayısı") * weight_kg
            
            total_vol_target = vol_coeff * weight_kg
            
            used_vol = (primene_vol + lipid_vol + nacl3_vol + kcl_vol + kpo4_vol + 
                        addamel_vol + cernevit_vol + dipeptiven_vol + omegaven_vol + 
                        mgso4_vol + ca_vol + heparin_vol)

            remaining_vol_for_dex = total_vol_target - used_vol
            if remaining_vol_for_dex < 0: remaining_vol_for_dex = 0

            # --- GLUKOZ SOLVER ---
            daily_glucose_g = (target_glucose * weight_kg * 1440) / 1000
            dex_types = [5, 10, 20, 30, 50]
            dex_vols = {5:0, 10:0, 20:0, 30:0, 50:0}
            
            if remaining_vol_for_dex > 0 and daily_glucose_g > 0:
                target_conc = (daily_glucose_g / remaining_vol_for_dex) * 100
                lower_dex = 5; upper_dex = 50
                
                for i in range(len(dex_types)-1):
                    if dex_types[i] <= target_conc <= dex_types[i+1]:
                        lower_dex = dex_types[i]
                        upper_dex = dex_types[i+1]
                        break
                
                if target_conc < 5: lower_dex=5; upper_dex=10
                if target_conc > 50: lower_dex=30; upper_dex=50

                denom = (upper_dex - lower_dex) / 100.0
                if denom == 0:
                    dex_vols[lower_dex] = remaining_vol_for_dex
                else:
                    v_upper = (daily_glucose_g - (remaining_vol_for_dex * lower_dex / 100.0)) / denom
                    v_lower = remaining_vol_for_dex - v_upper
                    if v_upper < 0: v_upper = 0; v_lower = remaining_vol_for_dex
                    if v_lower < 0: v_lower = 0; v_upper = remaining_vol_for_dex
                    dex_vols[upper_dex] = v_upper
                    dex_vols[lower_dex] = v_lower

            # --- SONUÇLARI SAKLA (1 Basamak yuvarlama) ---
            self.latest_results = {
                "ADI SOYADI": self.get_text("Adı Soyadı"),
                "TOTAL": round(total_vol_target, 1),
                "Tarih": self.get_text("Tarih"),
                "NaCl %0.9": 0,
                "%3 NaCl (serum sale)": round(nacl3_vol, 1),
                "KCL": round(kcl_vol, 1),
                "KPO4": round(kpo4_vol, 1),
                "ADDAMEL": round(addamel_vol, 1),
                "CERNEVİT": int(cernevit_vol),
                "LİPİT": round(lipid_vol, 1),
                "DİPEPTİVEN": round(dipeptiven_vol, 1),
                "OMEGAVEN": round(omegaven_vol, 1),
                "PRIMENE": round(primene_vol, 1),
                "DEXTROZ 5": round(dex_vols[5], 1),
                "DEXTROZ 10": round(dex_vols[10], 1),
                "DEXTROZ 20": round(dex_vols[20], 1),
                "DEXTROZ 30": round(dex_vols[30], 1),
                "DEXTROZ 50": round(dex_vols[50], 1),
                "MgSO4": round(mgso4_vol, 1),
                "Ca": round(ca_vol, 1),
                "HEPARİN": round(heparin_vol, 1)
            }
            
            for key, val in self.latest_results.items():
                self.result_labels[key].configure(text=str(val))
                if val == 0 or val == 0.0:
                     self.result_labels[key].configure(text_color="gray")
                else:
                     self.result_labels[key].configure(text_color="black")
            
            self.btn_save.configure(state="normal", fg_color="#1f6aa5")

        except Exception as e:
            messagebox.showerror("Hata", f"Hesaplama hatası: {str(e)}")

    def save_to_excel(self):
        if not self.latest_results:
            messagebox.showwarning("Uyarı", "Kaydedilecek veri bulunamadı.")
            return

        # --- YOL DÜZELTMESİ (Önemli Kısım) ---
        file_name = os.path.join(self.app_path, "TPN_Kayitlari.xlsx")
        
        headers = list(self.latest_results.keys())
        values = list(self.latest_results.values())

        try:
            if not os.path.exists(file_name):
                wb = Workbook()
                ws = wb.active
                ws.append(headers)
                ws.append(values)
                wb.save(file_name)
            else:
                wb = load_workbook(file_name)
                ws = wb.active
                ws.append(values)
                wb.save(file_name)
            
            messagebox.showinfo("Başarılı", f"Kayıt Eklendi!\nDosya Yeri: {file_name}")

        except Exception as e:
            messagebox.showerror("Hata", f"Kayıt Hatası:\n{e}\n\nLütfen EXE'yi masaüstünde bir klasöre çıkarıp deneyin.")

if __name__ == "__main__":
    app = TPNApp()
    app.mainloop()